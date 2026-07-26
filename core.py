# -*- coding: utf-8 -*-
"""三合一小工具核心逻辑: PDF转Excel / Excel公式 / AI简历。
与 UI 解耦, 便于测试与在 Streamlit Cloud 上部署。"""
import io
import json
import time
import os
import tempfile
import asyncio
import pandas as pd
import pdfplumber
import requests
from urllib.parse import quote
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

try:
    from PIL import Image, ImageDraw, ImageFont
    _PIL_AVAILABLE = True
except Exception:
    _PIL_AVAILABLE = False

try:
    import qrcode
    _QR_AVAILABLE = True
except Exception:
    _QR_AVAILABLE = False

try:
    import edge_tts
    _TTS_AVAILABLE = True
except Exception:
    _TTS_AVAILABLE = False

try:
    from pypdf import PdfReader, PdfWriter
    from reportlab.pdfgen import canvas as _rl_canvas
    from reportlab.lib.pagesizes import letter as _rl_letter
    from reportlab.pdfbase import pdfmetrics as _rl_pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont as _rl_TTFont
    _PDF_AVAILABLE = True
except Exception:
    _PDF_AVAILABLE = False


def _register_pdf_cjk_font():
    """为 reportlab 注册一个能显示中文的字体，跨平台找常见字体文件。"""
    if not _PDF_AVAILABLE:
        return None
    candidates = [
        # Windows
        ("MicrosoftYaHei", "C:/Windows/Fonts/msyh.ttc"),
        ("SimHei", "C:/Windows/Fonts/simhei.ttf"),
        ("SimSun", "C:/Windows/Fonts/simsun.ttc"),
        # Linux 常见
        ("WenQuanYiZenHei", "/usr/share/fonts/truetype/wqy/wqy-zenhei.ttc"),
        ("WenQuanYiMicroHei", "/usr/share/fonts/truetype/wqy/wqy-microhei.ttc"),
        ("NotoSansCJK", "/usr/share/fonts/truetype/noto/NotoSansCJK-Regular.ttc"),
        ("DejaVuSans", "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"),
    ]
    for name, path in candidates:
        try:
            _rl_pdfmetrics.registerFont(_rl_TTFont(name, path))
            return name
        except Exception:
            continue
    return None

# ===================== PDF 转 Excel =====================
def pdf_extract_tables(pdf_bytes):
    """返回 [(页码, 表序号, 二维表)] —— 二维表为 list[list[str|None]]。"""
    tables = []
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        for pi, page in enumerate(pdf.pages):
            found = page.extract_tables()
            for ti, tbl in enumerate(found):
                if tbl and len(tbl) > 0:
                    tables.append((pi + 1, ti + 1, tbl))
    return tables


def pdf_tables_to_excel(tables, merge_all=True, skip_empty=True):
    """把表格写成 Excel 字节流, 返回 (bytes, sheet_names)。"""
    output = io.BytesIO()
    if merge_all:
        all_rows = []
        for _pi, _ti, tbl in tables:
            for r in tbl:
                row = [str(c).strip() if c is not None else "" for c in r]
                if skip_empty and all(v == "" for v in row):
                    continue
                all_rows.append(row)
        maxc = max((len(r) for r in all_rows), default=1)
        df = pd.DataFrame([r + [""] * (maxc - len(r)) for r in all_rows])
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="合并表格", index=False, header=False)
            _style_sheet(writer.sheets["合并表格"])
        return output.getvalue(), ["合并表格"]
    else:
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            sheet_names = []
            for _idx, (pi, ti, tbl) in enumerate(tables):
                rows = []
                for r in tbl:
                    row = [str(c).strip() if c is not None else "" for c in r]
                    if skip_empty and all(v == "" for v in row):
                        continue
                    rows.append(row)
                if not rows:
                    continue
                maxc = max((len(r) for r in rows), default=1)
                df = pd.DataFrame([r + [""] * (maxc - len(r)) for r in rows])
                name = f"P{pi}_T{ti}"[:31]
                df.to_excel(writer, sheet_name=name, index=False, header=False)
                _style_sheet(writer.sheets[name])
                sheet_names.append(name)
        return output.getvalue(), sheet_names


def _style_sheet(ws):
    thin = Side(style="thin", color="D0D7E5")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill = PatternFill("solid", fgColor="2D6CDF")
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = hdr_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
    for row in ws.iter_rows():
        for c in row:
            c.border = border
            if c.row != 1:
                c.alignment = Alignment(vertical="center")
    for col in ws.columns:
        width = max((len(str(c.value)) for c in col if c.value is not None), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max(width + 2, 10), 40)


# ===================== 大模型调用 (公式 + 简历共用) =====================
PRESETS = {
    "DeepSeek (推荐·便宜)": {"base_url": "https://api.deepseek.com/v1", "model": "deepseek-chat"},
    "OpenAI": {"base_url": "https://api.openai.com/v1", "model": "gpt-4o-mini"},
    "通义千问": {"base_url": "https://dashscope.aliyuncs.com/compatible-mode/v1", "model": "qwen-plus"},
    "自定义": {"base_url": "", "model": ""},
}

FORMULA_MOCK = {
    "formula": "=SUM(A1:A10)/COUNT(A1:A10)",
    "explain": "先求 A1 到 A10 的总和，再除以个数，得到平均值。",
    "steps": [
        "用 SUM(A1:A10) 计算这一列的总和",
        "用 COUNT(A1:A10) 计算有多少个数字",
        "两者相除得到平均值（等价于 AVERAGE）",
    ],
    "example": "A1:A10 = 10,20,30 → SUM=60, COUNT=3 → 结果 20",
}

RESUME_MOCK = {
    "summary": "拥有 5 年互联网运营经验，擅长用户增长与内容策划，曾主导多个从 0 到 1 的项目，数据驱动决策，具备强执行力与跨团队协同能力。",
    "experience": [
        "● 主导某 APP 用户增长项目，3 个月内新增注册用户 12 万，留存率提升 18%",
        "● 搭建内容矩阵，公众号粉丝从 2 万增至 15 万，单篇最高阅读 50 万+",
        "● 设计裂变活动 SOP，获客成本降低 40%",
    ],
    "skills": ["用户增长", "内容运营", "数据分析", "活动策划", "SQL", "私域运营"],
    "education_tip": "建议补充 PMP 或 Google Analytics 证书以增强竞争力（如暂无相关学历背书）。",
    "ats_tips": [
        "使用标准职位名称（如'运营经理'）便于系统匹配",
        "技能区放置与 JD 高度重合的关键词",
        "避免表格/图片排版，纯文本更利于 ATS 解析",
    ],
}

FORMULA_SYSTEM = """你是一个 Excel 公式专家。用户会用自然语言描述他想要的计算，
请只返回一个 JSON 对象，不要包含任何多余文字或 markdown 代码块标记。
JSON 结构：
{
  "formula": "写在这里，例如 =SUM(A1:A10)",
  "explain": "用中文一句话解释这个公式在做什么",
  "steps": ["步骤1", "步骤2", "步骤3"],
  "example": "给出一个具体单元格的例子，例如 A1=10, A2=20 → 结果 30"
}
要求：formula 必须是可在 Excel / WPS 中直接使用的合法公式，以 = 开头。"""

RESUME_SYSTEM = """你是一位资深的中文简历优化顾问，熟悉 ATS（简历筛选系统）规则。
用户会提供：目标岗位、工作年限、以及一段原始经历/背景文字。
请只返回一个 JSON 对象，不要包含任何多余文字或 markdown 代码块标记。
JSON 结构：
{
  "summary": "3-4 句专业个人总结，突出核心竞争力与岗位匹配度",
  "experience": ["每条用 ● 开头，量化成果、用动词开头，例如 ● 主导XX项目，使转化率提升30%"],
  "skills": ["技能关键词1","技能关键词2"],
  "education_tip": "针对其背景的学历/证书建议（若用户未提供则给通用建议）",
  "ats_tips": ["ATS 优化建议1","ATS 优化建议2"]
}
要求：全部使用简体中文，专业、具体、可量化，避免空话。"""


def _call_llm(system_prompt, user_msg, api_key, base_url, model):
    payload = {
        "model": model,
        "messages": [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_msg},
        ],
        "temperature": 0.3,
        "response_format": {"type": "json_object"},
    }
    headers = {"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"}
    try:
        r = requests.post(
            base_url.rstrip("/") + "/chat/completions",
            headers=headers, json=payload, timeout=30,
        )
        r.raise_for_status()
        return json.loads(r.json()["choices"][0]["message"]["content"])
    except requests.exceptions.HTTPError:
        # 部分模型不支持 response_format，降级重试一次
        payload.pop("response_format", None)
        r2 = requests.post(
            base_url.rstrip("/") + "/chat/completions",
            headers=headers, json=payload, timeout=30,
        )
        r2.raise_for_status()
        return json.loads(r2.json()["choices"][0]["message"]["content"])


def formula_generate(prompt, api_key, base_url, model, mock=False):
    """调用大模型生成 Excel 公式。mock=True 或缺少 key 时返回演示数据。"""
    if mock or not api_key or not base_url:
        return FORMULA_MOCK
    return _call_llm(FORMULA_SYSTEM, prompt, api_key, base_url, model)


def resume_generate(target_role, years, background, api_key, base_url, model, mock=False):
    """调用大模型生成优化简历。mock=True 或缺少 key 时返回演示数据。"""
    if mock or not api_key or not base_url:
        return RESUME_MOCK
    user_msg = f"目标岗位：{target_role}\n工作年限：{years}\n原始背景/经历：\n{background}"
    return _call_llm(RESUME_SYSTEM, user_msg, api_key, base_url, model)


# ===================== 签证要求聚合 (信息差版) =====================
import json as _json
import os as _os

def load_visa_data():
    """读取签证数据集, 返回 (meta, list[dict])。"""
    path = _os.path.join(_os.path.dirname(__file__), "visa_data.json")
    with open(path, "r", encoding="utf-8") as f:
        data = _json.load(f)
    return data.get("_meta", {}), data.get("countries", [])


def visa_search(keyword="", region="全部", policy="全部"):
    """按关键词(国家名) / 地区 / 签证类型筛选。返回 list[dict]。"""
    _meta, countries = load_visa_data()
    kw = keyword.strip().lower()
    result = []
    for c in countries:
        if kw and kw not in c["country"].lower() and kw not in c.get("region", "").lower():
            continue
        if region != "全部" and c.get("region") != region:
            continue
        if policy != "全部" and c.get("policy") != policy:
            continue
        result.append(c)
    return result


def visa_regions():
    _meta, countries = load_visa_data()
    rs = sorted({c.get("region", "其他") for c in countries})
    return ["全部"] + rs


def visa_policies():
    return ["全部", "免签", "免签/落地签", "落地签", "落地签/电子签", "电子签",
            "电子旅行授权", "电子签/持美签免签", "需提前办理"]


# ===================== 数据可视化 =====================
def viz_load_data(raw_bytes, filename):
    """从上传的 CSV/Excel 字节加载为 DataFrame。"""
    name = (filename or "").lower()
    if name.endswith((".xlsx", ".xls")):
        return pd.read_excel(io.BytesIO(raw_bytes))
    last_err = None
    for enc in ("utf-8-sig", "utf-8", "gbk", "gb18030", "latin-1"):
        try:
            return pd.read_csv(io.BytesIO(raw_bytes), encoding=enc)
        except Exception as e:
            last_err = e
    raise ValueError(f"无法解析该文件（试试 UTF-8 或 GBK 编码的 CSV）：{last_err}")


def viz_is_numeric(series):
    """判断某列是否可当作数值（原生数值，或多数能转数值）。"""
    if pd.api.types.is_numeric_dtype(series):
        return True
    return pd.to_numeric(series, errors="coerce").notna().mean() > 0.5


def viz_make_figure(df, chart_type, x_col, y_cols, title="", agg="不聚合", top_n=0):
    """返回 plotly.graph_objects.Figure。

    chart_type: 柱状图 / 折线图 / 饼图 / 散点图
    agg: 不聚合 / 求和 / 平均 / 计数（x 为分类列、y 为数值列时生效）
    top_n: >0 时按 y 第一列取前 N 行（绘图前排序）
    """
    import plotly.graph_objects as go
    work = df.copy()
    if agg != "不聚合" and x_col and y_cols:
        try:
            if agg == "求和":
                work = work.groupby(x_col, as_index=False)[y_cols].sum(numeric_only=True)
            elif agg == "平均":
                work = work.groupby(x_col, as_index=False)[y_cols].mean(numeric_only=True)
            elif agg == "计数":
                work = work.groupby(x_col, as_index=False)[y_cols].count()
        except Exception:
            pass
    if top_n and top_n > 0 and y_cols:
        y0 = y_cols[0]
        try:
            work = work.sort_values(y0, ascending=False).head(top_n)
        except Exception:
            pass

    fig = go.Figure()
    if chart_type == "柱状图":
        for y in y_cols:
            fig.add_bar(name=str(y), x=work[x_col], y=work[y])
        fig.update_layout(barmode="group")
    elif chart_type == "折线图":
        for y in y_cols:
            fig.add_trace(go.Scatter(x=work[x_col], y=work[y], mode="lines+markers", name=str(y)))
    elif chart_type == "饼图":
        y = y_cols[0] if y_cols else None
        fig.add_trace(go.Pie(labels=work[x_col], values=work[y], textinfo="label+percent"))
    elif chart_type == "散点图":
        y = y_cols[0] if y_cols else None
        if y:
            fig.add_trace(go.Scatter(x=work[x_col], y=work[y], mode="markers", name=str(y)))
    fig.update_layout(
        title=title or "数据可视化",
        template="plotly_white",
        font=dict(family="Microsoft YaHei, PingFang SC, sans-serif", size=13),
        margin=dict(l=40, r=20, t=50, b=40),
        legend=dict(orientation="h", y=-0.2),
    )
    return fig


# ===================== CSV → Airtable 导入 =====================
def airtable_list_tables(token, base_id):
    """拉取 base 下的表结构。返回 list[{name, id, fields:[{name,type}]}]。"""
    url = f"https://api.airtable.com/v0/meta/bases/{base_id}/tables"
    headers = {"Authorization": f"Bearer {token}"}
    r = requests.get(url, headers=headers, timeout=20)
    r.raise_for_status()
    data = r.json()
    out = []
    for t in data.get("tables", []):
        out.append({
            "name": t.get("name"),
            "id": t.get("id"),
            "fields": [{"name": f.get("name"), "type": f.get("type")} for f in t.get("fields", [])],
        })
    return out


def airtable_import(token, base_id, table_name, records, dry_run=False, batch=10):
    """把已映射好的 records 批量写入 Airtable。

    records: list[dict]，每个 dict 是 {字段名: 值}。
    dry_run=True 时只模拟、不真实写入（用于演示流程）。
    返回 dict: {success, failed, errors, dry}。
    """
    if dry_run:
        return {"success": len(records), "failed": 0, "errors": [], "dry": True}
    url = f"https://api.airtable.com/v0/{base_id}/{quote(table_name)}"
    headers = {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}
    ok = 0
    failed = 0
    errors = []
    for i in range(0, len(records), batch):
        chunk = records[i:i + batch]
        payload = {"records": [{"fields": rec} for rec in chunk], "typecast": True}
        try:
            r = requests.post(url, headers=headers, json=payload, timeout=30)
            if r.status_code in (200, 201):
                ok += len(chunk)
            else:
                failed += len(chunk)
                errors.append(f"批次 {i // batch + 1}: HTTP {r.status_code} {r.text[:200]}")
        except Exception as e:
            failed += len(chunk)
            errors.append(f"批次 {i // batch + 1}: {e}")
        time.sleep(0.25)  # 尊重 Airtable 限速（5 次/秒）
    return {"success": ok, "failed": failed, "errors": errors}


# ===================== AI 配图 / 海报生成器 =====================
POSTER_THEMES = {
    "商务蓝": {"bg": "linear-gradient(135deg,#1E3A8A 0%,#2D6CDF 100%)", "accent": "#7CFC9B",
               "text": "#FFFFFF", "sub": "#D6E2FF", "deco": "rgba(255,255,255,0.12)"},
    "活力橙": {"bg": "linear-gradient(135deg,#FF6A00 0%,#FFB347 100%)", "accent": "#FFF3E0",
               "text": "#FFFFFF", "sub": "#FFE9CC", "deco": "rgba(255,255,255,0.16)"},
    "极简黑": {"bg": "linear-gradient(135deg,#111827 0%,#1F2937 100%)", "accent": "#FBBF24",
               "text": "#FFFFFF", "sub": "#9CA3AF", "deco": "rgba(255,255,255,0.10)"},
    "清新绿": {"bg": "linear-gradient(135deg,#065F46 0%,#10B981 100%)", "accent": "#D1FAE5",
               "text": "#FFFFFF", "sub": "#CFFAF0", "deco": "rgba(255,255,255,0.14)"},
    "少女粉": {"bg": "linear-gradient(135deg,#DB2777 0%,#F9A8D4 100%)", "accent": "#FFF0F6",
               "text": "#FFFFFF", "sub": "#FCE7F3", "deco": "rgba(255,255,255,0.18)"},
    "高级紫": {"bg": "linear-gradient(135deg,#4C1D95 0%,#7C3AED 100%)", "accent": "#EDE9FE",
               "text": "#FFFFFF", "sub": "#DDD6FE", "deco": "rgba(255,255,255,0.12)"},
    # ===== 第二批模板（共 12 套）=====
    "科技青": {"bg": "linear-gradient(135deg,#0F766E 0%,#06B6D4 100%)", "accent": "#A7F3D0",
               "text": "#FFFFFF", "sub": "#CCFBF1", "deco": "rgba(255,255,255,0.12)"},
    "暖阳橙金": {"bg": "linear-gradient(135deg,#B45309 0%,#F59E0B 100%)", "accent": "#FEF3C7",
                 "text": "#FFFFFF", "sub": "#FDE68A", "deco": "rgba(255,255,255,0.14)"},
    "莫兰迪灰": {"bg": "linear-gradient(135deg,#374151 0%,#9CA3AF 100%)", "accent": "#F3F4F6",
                 "text": "#FFFFFF", "sub": "#E5E7EB", "deco": "rgba(255,255,255,0.10)"},
    "赛博粉": {"bg": "linear-gradient(135deg,#831843 0%,#DB2777 100%)", "accent": "#FBCFE8",
               "text": "#FFFFFF", "sub": "#FCE7F3", "deco": "rgba(255,255,255,0.14)"},
    "深海蓝": {"bg": "linear-gradient(135deg,#0C4A6E 0%,#0369A1 100%)", "accent": "#BAE6FD",
               "text": "#FFFFFF", "sub": "#E0F2FE", "deco": "rgba(255,255,255,0.12)"},
    "咖啡棕": {"bg": "linear-gradient(135deg,#451A03 0%,#92400E 100%)", "accent": "#FED7AA",
               "text": "#FFFFFF", "sub": "#FFEDD5", "deco": "rgba(255,255,255,0.10)"},
}

POSTER_SIZES = {
    "小红书 3:4": (1080, 1350),
    "公众号封面 2.35:1": (900, 383),
    "YouTube缩略图 16:9": (1280, 720),
    "朋友圈方图 1:1": (1080, 1080),
    "抖音·视频号 9:16": (1080, 1920),
}


def poster_themes():
    return list(POSTER_THEMES.keys())


def poster_sizes():
    return list(POSTER_SIZES.keys())


def generate_poster_html(title, subtitle, theme, size, footer, tags):
    """生成一张可直接在浏览器渲染的配图 HTML（中文用系统字体，无需字体文件）。"""
    th = POSTER_THEMES.get(theme, POSTER_THEMES["商务蓝"])
    w, h = POSTER_SIZES.get(size, (1080, 1350))
    # 注意：此处不能用双引号包裹字体名，因为 font_stack 会被嵌入
    # style="..." 双引号属性中，内层双引号会导致 HTML 解析断裂、CSS 原样泄露为可见文字。
    font_stack = "'PingFang SC','Microsoft YaHei','Hiragino Sans GB',sans-serif"

    tag_html = ""
    if tags:
        pills = "".join(
            f'<span style="display:inline-block;background:{th["accent"]};color:#111827;'
            f'font-weight:800;padding:8px 18px;border-radius:999px;margin:5px;'
            f'font-size:{int(h*0.03)}px;font-family:{font_stack};">#{t}</span>'
            for t in tags
        )
        tag_html = f'<div style="margin-top:{int(h*0.05)}px;">{pills}</div>'

    footer_html = ""
    if footer:
        footer_html = (
            f'<div style="position:absolute;bottom:{int(h*0.055)}px;left:0;right:0;'
            f'text-align:center;color:{th["sub"]};font-size:{int(h*0.032)}px;'
            f'font-family:{font_stack};letter-spacing:1px;">{footer}</div>'
        )

    sub_html = ""
    if subtitle:
        sub_html = (
            f'<div style="color:{th["accent"]};font-weight:800;'
            f'font-size:{int(h*0.05)}px;letter-spacing:3px;margin-bottom:{int(h*0.035)}px;'
            f'font-family:{font_stack};">{subtitle}</div>'
        )

    html = f'''<div style="position:relative;width:{w}px;height:{h}px;background:{th["bg"]};overflow:hidden;font-family:{font_stack};">
  <div style="position:absolute;top:-{int(w*0.25)}px;right:-{int(w*0.2)}px;width:{int(w*0.7)}px;height:{int(w*0.7)}px;border-radius:50%;background:{th["deco"]};"></div>
  <div style="position:absolute;bottom:-{int(w*0.18)}px;left:-{int(w*0.15)}px;width:{int(w*0.55)}px;height:{int(w*0.55)}px;border-radius:50%;background:{th["deco"]};"></div>
  <div style="position:absolute;inset:0;display:flex;flex-direction:column;justify-content:center;align-items:center;padding:{int(w*0.1)}px;text-align:center;box-sizing:border-box;">
    {sub_html}
    <div style="color:{th["text"]};font-weight:900;font-size:{int(h*0.09)}px;line-height:1.18;font-family:{font_stack};">{title}</div>
    {tag_html}
  </div>
  {footer_html}
</div>'''
    return html


def poster_preview_html(poster_html, w, h, max_w=480):
    """把整张海报按比例缩小放进预览框，保持原比例。"""
    scale = round(max_w / w, 4)
    return (
        f'<div style="width:{int(w*scale)}px;height:{int(h*scale)}px;overflow:hidden;'
        f'margin:0 auto;border-radius:14px;box-shadow:0 6px 24px rgba(0,0,0,0.12);">'
        f'<div style="transform:scale({scale});transform-origin:top left;width:{w}px;height:{h}px;">'
        f'{poster_html}</div></div>'
    )


POSTER_MOCK = {
    "title": "3 个被低估的赚钱小工具",
    "subtitle": "信息差 · 普通人也能上手",
    "tags": ["副业", "AI工具", "信息差"],
}

POSTER_SYSTEM = """你是一个社媒配图文案专家，擅长写抓眼球的中文标题。
用户会给他想做的内容主题。请只返回一个 JSON 对象，不要包含多余文字或 markdown 标记。
JSON 结构：
{
  "title": "一句抓眼球的标题（12-20字，用换行符\\n分成1-2行更醒目）",
  "subtitle": "一行小字标签（如 信息差·副业 这类，8字内）",
  "tags": ["标签1","标签2","标签3"]
}
要求：全部简体中文，标题要有钩子感（数字/反差/痛点），避免空话。"""


def poster_copy(topic, api_key, base_url, model, mock=False):
    """调用大模型为配图生成文案。缺 Key 或 mock 时返回演示文案。"""
    if mock or not api_key or not base_url:
        return POSTER_MOCK
    user_msg = f"内容主题：{topic}\n请为它写一张社媒配图的标题与标签。"
    return _call_llm(POSTER_SYSTEM, user_msg, api_key, base_url, model)


# ===================== 图片处理工具箱 =====================
def image_process(raw_bytes, *, fmt="保持原格式", quality=85, max_edge=0, watermark=None):
    """处理单张图片，返回 (out_bytes, ext, w, h)。

    fmt: 保持原格式 / JPG / PNG
    quality: 1-95（仅对 JPG / 原格式为 JPEG 时生效）
    max_edge: 最长边像素，0 表示不缩放
    watermark: 水印文字（None / 空则不加水印）
    """
    if not _PIL_AVAILABLE:
        raise EnvironmentError("图片处理需要 Pillow 库，请确认 requirements.txt 中包含 Pillow")

    img = Image.open(io.BytesIO(raw_bytes))
    orig_format = (img.format or "JPEG").upper()
    img = img.copy()  # 避免 "image file is closed" 问题

    # 选输出格式 + 透明通道处理
    if fmt == "JPG":
        if img.mode in ("RGBA", "LA", "P"):
            bg = Image.new("RGB", img.size, (255, 255, 255))
            if img.mode == "P":
                img = img.convert("RGBA")
            mask = img.split()[-1] if img.mode in ("RGBA", "LA") else None
            bg.paste(img, mask=mask)
            img = bg
        else:
            img = img.convert("RGB")
        save_fmt = "JPEG"
    elif fmt == "PNG":
        save_fmt = "PNG"
    else:  # 保持原格式
        # 本环境 PIL 未编译 WebP 编码，原图为 WebP 时退回 PNG
        save_fmt = orig_format if orig_format in ("JPEG", "PNG", "BMP", "GIF") else "PNG"
        if save_fmt == "JPEG":
            img = img.convert("RGB")

    # 缩放：只缩小不放大
    if max_edge and max_edge > 0:
        w, h = img.size
        longest = max(w, h)
        if longest > max_edge:
            scale = max_edge / float(longest)
            img = img.resize((int(round(w * scale)), int(round(h * scale))),
                             Image.LANCZOS)

    # 水印
    if watermark and watermark.strip():
        img = _add_watermark(img, watermark.strip())

    out = io.BytesIO()
    if save_fmt == "JPEG":
        img.save(out, format="JPEG", quality=int(quality), optimize=True)
    elif save_fmt == "PNG":
        img.save(out, format="PNG", optimize=True)
    else:
        img.save(out, format=save_fmt, optimize=True)
    return out.getvalue(), save_fmt.lower(), img.size[0], img.size[1]


def _add_watermark(img, text):
    """在右下角加半透明文字水印（带描边便于在任何底色上可读）。"""
    if img.mode not in ("RGB", "RGBA"):
        img = img.convert("RGB")
    draw = ImageDraw.Draw(img, "RGBA")
    fs = max(16, int(min(img.size) / 28))
    try:
        font = ImageFont.truetype("arial.ttf", fs)
    except Exception:
        try:
            font = ImageFont.truetype("C:/Windows/Fonts/msyh.ttc", fs)
        except Exception:
            font = ImageFont.load_default()
    bbox = draw.textbbox((0, 0), text, font=font)
    tw, th = bbox[2] - bbox[0], bbox[3] - bbox[1]
    x = img.width - tw - max(8, int(img.width * 0.03))
    y = img.height - th - max(8, int(img.height * 0.03))
    # 先画暗色描边/底，再画亮色字，保证对比度
    draw.text((x + 2, y + 2), text, font=font, fill=(0, 0, 0, 150))
    draw.text((x, y), text, font=font, fill=(255, 255, 255, 210))
    return img


def image_tool_formats():
    """返回可选的输出格式列表。"""
    return ["保持原格式", "JPG", "PNG"]


# ===================== 二维码生成器 =====================
def qr_generate(text, *, box_size=10, border=4, error_correction="M",
                fg_color="#000000", bg_color="#FFFFFF", logo_bytes=None,
                logo_scale=0.20):
    """生成二维码，返回 (png_bytes, size_px)。

    error_correction: L(7%) / M(15%) / Q(25%) / H(30%) 容错率
    logo_bytes: 可选的中心 Logo 图片字节（PNG/JPG）
    logo_scale: Logo 占二维码边长的比例（0-0.3）
                  加 Logo 时会自动限制：L≤0.06 / M≤0.12 / Q≤0.18 / H≤0.25，
                  超限时自动升级容错率保证可扫。
    """
    if not _QR_AVAILABLE:
        raise EnvironmentError("二维码需要 qrcode 库，请确认 requirements.txt 中包含 qrcode")
    if not text or not text.strip():
        raise ValueError("二维码内容不能为空")

    # 各容错率允许的最大 Logo 比例（留安全余量）
    ec_max_logo = {"L": 0.06, "M": 0.12, "Q": 0.18, "H": 0.25}
    ec_order = ["L", "M", "Q", "H"]

    # 如果有 Logo 且当前容错率放不下 → 自动升级
    if logo_bytes and float(logo_scale) > ec_max_logo.get(error_correction, 0):
        for higher in ec_order:
            if ec_order.index(higher) > ec_order.index(error_correction):
                if float(logo_scale) <= ec_max_logo.get(higher, 0.30):
                    error_correction = higher
                    break

    # 最终安全上限（不超过当前容错率可恢复范围）
    safe_max = ec_max_logo.get(error_correction, 0.25)
    actual_scale = min(float(logo_scale), safe_max)

    ec_map = {
        "L": qrcode.constants.ERROR_CORRECT_L,
        "M": qrcode.constants.ERROR_CORRECT_M,
        "Q": qrcode.constants.ERROR_CORRECT_Q,
        "H": qrcode.constants.ERROR_CORRECT_H,
    }
    qr = qrcode.QRCode(
        version=None,
        error_correction=ec_map[error_correction],
        box_size=box_size,
        border=border,
    )
    qr.add_data(text.strip())
    qr.make(fit=True)

    if not _PIL_AVAILABLE:
        img = qr.make_image(fill_color="black", back_color="white")
    else:
        img = qr.make_image(fill_color=fg_color, back_color=bg_color).convert("RGB")

    # 中心 Logo（需 PIL）—— 用安全缩放比例
    if logo_bytes and _PIL_AVAILABLE and actual_scale > 0:
        try:
            logo = Image.open(io.BytesIO(logo_bytes)).convert("RGBA")
            qr_w, qr_h = img.size
            logo_size = int(min(qr_w, qr_h) * actual_scale)
            logo_size = max(16, logo_size)
            logo = logo.resize((logo_size, logo_size), Image.LANCZOS)
            pad = max(4, int(logo_size * 0.10))
            plate = Image.new("RGBA", (logo_size + pad * 2, logo_size + pad * 2),
                              (255, 255, 255, 255))
            plate.paste(logo, (pad, pad), logo)
            px, py = (qr_w - plate.width) // 2, (qr_h - plate.height) // 2
            img.paste(plate, (px, py), plate)
        except Exception:
            pass

    out = io.BytesIO()
    img.save(out, format="PNG")
    return out.getvalue(), img.size[0]


# ===================== 文字转语音 TTS =====================
TTS_VOICES = {
    "晓晓（女·甜美）": "zh-CN-XiaoxiaoNeural",
    "云希（男·沉稳）": "zh-CN-YunxiNeural",
    "云剑（男·专业）": "zh-CN-YunjianNeural",
    "晓伊（女·温柔）": "zh-CN-XiaoyiNeural",
    "云夏（男·活力）": "zh-CN-YunxiaNeural",
    "晓睿（女·活泼）": "zh-CN-XiaoruiNeural",
    "云扬（男·新闻）": "zh-CN-YunyangNeural",
}


def tts_voices():
    """返回可选音色的中文标签列表。"""
    return list(TTS_VOICES.keys())


def tts_generate(text, voice_label="晓晓（女·甜美）", rate="+0%", pitch="+0Hz"):
    """调用微软免费 TTS（免 Key）生成语音，返回 MP3 字节流。

    rate: 语速，如 +20% / -10%；pitch: 音调，如 +10Hz / -5Hz。
    """
    if not _TTS_AVAILABLE:
        raise EnvironmentError("文字转语音需要 edge-tts，请确认 requirements.txt 中包含 edge-tts")
    if not text or not text.strip():
        raise ValueError("文字内容不能为空")

    short = TTS_VOICES.get(voice_label, "zh-CN-XiaoxiaoNeural")
    tmp = tempfile.NamedTemporaryFile(suffix=".mp3", delete=False)
    tmp.close()
    try:
        communicate = edge_tts.Communicate(text.strip(), short, rate=rate, pitch=pitch)
        loop = asyncio.new_event_loop()
        try:
            asyncio.set_event_loop(loop)
            loop.run_until_complete(communicate.save(tmp.name))
        finally:
            loop.close()
        with open(tmp.name, "rb") as f:
            data = f.read()
        return data
    finally:
        try:
            os.unlink(tmp.name)
        except Exception:
            pass


# ===================== 多平台文案改写 (一稿多发) =====================
REWRITE_PLATFORMS = ["小红书", "公众号", "抖音/视频号", "微博", "知乎"]


def rewrite_platforms():
    """返回可选的目标平台列表。"""
    return REWRITE_PLATFORMS


REWRITE_SYSTEM = """你是一个资深的新媒体多平台文案改写专家。
当用户给出一段原始文案和一个目标平台时，请把文案改写成该平台最合适的风格与格式。

各平台风格要点：
- 小红书：大量 emoji、口语化、第一人称种草、分段清晰、结尾加 3-5 个 #话题标签（如 #副业 #信息差 #小工具）、带互动引导（如"评论区告诉我"）
- 公众号：标题有吸引力但不浮夸、开篇抛观点/痛点、结构清晰有小标题、语气专业可信、适合 300-800 字
- 抖音/视频号：强钩子开头（前 3 秒抓人）、短句为主、口语化带节奏感、适合口播、可加 2-3 个 emoji
- 微博：短平快、140 字内最佳、带话题 #、有互动感（提问/投票）
- 知乎：理性专业、有逻辑层次、可引用数据或案例、结尾可引导关注，语气克制

只返回改写后的文案正文，不要包含任何解释或 markdown 代码块标记。"""


REWRITE_MOCK = {
    "小红书": "✨姐妹们！今天挖到一个超好用的小工具站🔧\n\n里面整整 12 款工具，从 PDF 转 Excel 到 AI 配图、二维码、文字转语音全都有，关键还全免费！\n\n我自己的使用感受：界面清爽、不卡顿，上传文件本地处理很安心🛡️\n\n#副业 #信息差 #小工具推荐 #效率神器 #AI工具",
    "公众号": "你缺的不是一个工具，而是一套能赚钱的小工具组合。\n\n今天想认真安利一个我最近在用的「小工具站」——12 款实用工具聚合在一起，覆盖 PDF 处理、Excel 公式、AI 简历、配图、二维码、语音等场景。\n\n为什么值得用？\n1. 全免费，打开就能用\n2. 本地处理，数据不上传\n3. 对标海外成功案例，功能经过验证\n\n如果你也在做副业或内容创作，建议收藏。",
    "抖音/视频号": "🔥别再到处找工具了！这一个网站全搞定！\n\n12 款神器，PDF 转 Excel、AI 写简历、一键配图、语音生成…全免费！\n\n而且本地处理，隐私不怕泄露！\n\n赶紧收藏，错过真的亏！",
    "微博": "挖到一个免费小工具站，12 款实用工具聚合，PDF/Excel/配图/二维码/语音全有，本地处理很安心。做副业和内容创作的值得收藏👉 #效率工具 #副业 #信息差",
    "知乎": "如何看待「小工具聚合站」这种低成本创业形态？\n\n近期体验了一款聚合 12 款实用工具的小工具站，谈几点观察：\n\n1. 模式上，它把海外验证过的单点工具（如 PDF 处理、AI 简历）打包成一站，降低用户决策成本；\n2. 隐私上采用本地处理，符合国内用户对数据安全的诉求；\n3. 变现路径清晰——免费引流，增值服务转化。\n\n这类「轻资产、强需求」的产品，值得关注。",
}


def _call_llm_text(system_prompt, user_msg, api_key, base_url, model):
    """调用大模型返回纯文本（非 JSON）。用于文案改写等自由文本场景。"""
    payload = {
        "model": model,
        "messages": [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_msg},
        ],
        "temperature": 0.7,
    }
    headers = {"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"}
    r = requests.post(base_url.rstrip("/") + "/chat/completions",
                      headers=headers, json=payload, timeout=30)
    r.raise_for_status()
    return r.json()["choices"][0]["message"]["content"].strip()


def rewrite_copy(text, platform, api_key, base_url, model, mock=False):
    """按目标平台风格改写文案。缺 Key 或 mock 时返回演示改写。"""
    platform = platform or "小红书"
    text = (text or "").strip()
    if not text:
        raise ValueError("请先输入要改写的原始文案")
    if mock or not api_key or not base_url:
        return REWRITE_MOCK.get(platform, text)
    user_msg = f"目标平台：{platform}\n原始文案：\n{text}"
    return _call_llm_text(REWRITE_SYSTEM, user_msg, api_key, base_url, model)


# ===================== PDF 合并 / 加水印 =====================
def pdf_merge(pdf_list):
    """合并多个 PDF 字节流，返回 (合并后字节, 页数)。pdf_list: [(文件名, 字节)]。"""
    if not _PDF_AVAILABLE:
        raise EnvironmentError("PDF 处理需要 pypdf 库，请确认 requirements.txt 中包含 pypdf")
    if not pdf_list:
        raise ValueError("请先上传至少一个 PDF 文件")
    writer = PdfWriter()
    for _name, raw in pdf_list:
        reader = PdfReader(io.BytesIO(raw))
        writer.append(reader)
    out = io.BytesIO()
    writer.write(out)
    out.seek(0)
    n = len(PdfReader(out).pages)
    return out.getvalue(), n


def pdf_add_watermark(pdf_bytes, text, *, angle=-30, opacity=0.18,
                      color=(128, 128, 128), font_size=48, density=1):
    """给 PDF 每页加平铺文字水印，返回 (加水印后字节, 页数)。
    density: 1=稀疏, 2=中等, 3=密集。
    """
    if not _PDF_AVAILABLE:
        raise EnvironmentError("PDF 处理需要 pypdf 库，请确认 requirements.txt 中包含 pypdf")
    if not pdf_bytes:
        raise ValueError("请先上传 PDF 文件")
    if not text or not text.strip():
        raise ValueError("水印文字不能为空")
    text = text.strip()

    reader = PdfReader(io.BytesIO(pdf_bytes))
    n = len(reader.pages)
    if n == 0:
        raise ValueError("PDF 没有页面")

    # 用 reportlab 生成一张与首页同尺寸的水印图层
    first_page = reader.pages[0]
    box = first_page.mediabox
    w = float(box.width)
    h = float(box.height)

    # 平铺间距随密度变化
    step = {1: 320, 2: 240, 3: 170}.get(density, 240)
    r, g, b = color

    # 注册中文字体；找不到则 fallback 到 Helvetica-Bold（中文会显示异常）
    font_name = _register_pdf_cjk_font() or "Helvetica-Bold"

    wm_buf = io.BytesIO()
    c = _rl_canvas.Canvas(wm_buf, pagesize=(w, h))
    c.setFillColorRGB(r / 255.0, g / 255.0, b / 255.0, alpha=opacity)
    c.setFont(font_name, font_size)
    c.rotate(angle)
    # 旋转后坐标空间变了，用足够大的网格覆盖
    span = int((w + h) * 1.5)
    x = -span
    while x < span:
        y = -span
        while y < span:
            c.drawString(x, y, text)
            y += step
        x += step + font_size
    c.save()
    wm_buf.seek(0)
    wm_reader = PdfReader(wm_buf)
    wm_page = wm_reader.pages[0]

    writer = PdfWriter()
    for page in reader.pages:
        page.merge_page(wm_page)
        writer.add_page(page)
    out = io.BytesIO()
    writer.write(out)
    return out.getvalue(), n


def pdf_page_count(pdf_bytes):
    """返回 PDF 页数。"""
    if not _PDF_AVAILABLE:
        return None
    try:
        return len(PdfReader(io.BytesIO(pdf_bytes)).pages)
    except Exception:
        return None


# ===================== 短链接生成 =====================
SHORTEN_SERVICES = {
    # tinyurl / cleanuri 对微信/QQ 域名更友好；is.gd/v.gd 近期对很多域名会黑名单或数据库失败
    "tinyurl.com (兼容微信/QQ)": {
        "endpoint": "https://tinyurl.com/api-create.php",
        "method": "GET",
        "params": lambda url: {"url": url},
        "extract": lambda resp: resp.text.strip(),
    },
    "cleanuri.com (兼容微信/QQ)": {
        "endpoint": "https://cleanuri.com/api/v1/shorten",
        "method": "POST",
        "data": lambda url: {"url": url},
        "extract": lambda resp: resp.json().get("result_url", "").strip(),
    },
    "is.gd (国际通用)": {
        "endpoint": "https://is.gd/create.php",
        "method": "GET",
        "params": lambda url: {"format": "simple", "url": url},
        "extract": lambda resp: resp.text.strip(),
    },
    "v.gd (国际通用)": {
        "endpoint": "https://v.gd/create.php",
        "method": "GET",
        "params": lambda url: {"format": "simple", "url": url},
        "extract": lambda resp: resp.text.strip(),
    },
}


def shorten_services():
    return list(SHORTEN_SERVICES.keys())


def shorten_url(url, service="is.gd (推荐·稳定)", timeout=15):
    """调用免费短链服务把长链接缩短，返回短网址字符串。
    支持 is.gd / v.gd / tinyurl / cleanuri，免 Key、免注册。
    微信/QQ 域名可能被 is.gd/v.gd 黑名单拦截，可切 tinyurl/cleanuri。
    """
    url = (url or "").strip()
    if not url:
        raise ValueError("请先输入要缩短的网址")
    if not (url.startswith("http://") or url.startswith("https://")):
        raise ValueError("网址需以 http:// 或 https:// 开头")
    cfg = SHORTEN_SERVICES.get(service, SHORTEN_SERVICES["is.gd (推荐·稳定)"])
    try:
        kwargs = {"timeout": timeout}
        if "params" in cfg:
            kwargs["params"] = cfg["params"](url)
        if "data" in cfg:
            kwargs["data"] = cfg["data"](url)
        resp = requests.request(cfg["method"], cfg["endpoint"], **kwargs)
    except Exception as e:
        raise RuntimeError(f"请求短链服务失败（可能网络受限）：{e}")
    if resp.status_code != 200:
        raise RuntimeError(f"短链服务返回错误状态码 {resp.status_code}")
    text = cfg["extract"](resp)
    # is.gd / v.gd / tinyurl 失败时会返回 "Error: ..." 文本；cleanuri 失败会无 result_url
    if (not text) or (text.lower().startswith("error")):
        detail = text or "未返回短链结果"
        if "blacklist" in detail.lower():
            raise ValueError(f"该域名被此服务列入黑名单，请换用 tinyurl.com / cleanuri.com：{detail}")
        raise ValueError(f"短链生成失败：{detail}")
    return text


# ===================== 图片智能去背景 =====================
try:
    from rembg import remove as _rembg_remove
    _REMBG_AVAILABLE = True
except Exception:
    _REMBG_AVAILABLE = False


def remove_bg(image_bytes, *, bg_type="transparent", bg_color=(255, 255, 255)):
    """用 rembg 移除图片背景。
    bg_type: transparent(透明 PNG) / white(白底) / color(自定义纯色底)。
    返回 PNG 字节。
    """
    if not _PIL_AVAILABLE:
        raise EnvironmentError("图片处理需要 Pillow")
    if not _REMBG_AVAILABLE:
        raise EnvironmentError("智能去背景需要 rembg 库（含 onnxruntime），请确认 requirements.txt 包含 rembg")
    if not image_bytes:
        raise ValueError("请先上传图片")
    try:
        img = Image.open(io.BytesIO(image_bytes)).convert("RGB")
    except Exception:
        raise ValueError("图片无法识别，请换一张 PNG/JPG")
    # rembg 返回 RGBA
    cut = _rembg_remove(img)
    if bg_type == "transparent":
        final = cut
    else:
        if bg_type == "color":
            rgb = tuple(int(c) for c in bg_color[:3])
        else:
            rgb = (255, 255, 255)
        base = Image.new("RGBA", cut.size, rgb + (255,))
        final = Image.alpha_composite(base, cut)
    buf = io.BytesIO()
    final.save(buf, "PNG")
    return buf.getvalue()


